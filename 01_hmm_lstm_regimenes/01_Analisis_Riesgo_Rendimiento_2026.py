# ============================================================
# 05_HMM_Hibrido.py
# Análisis híbrido HMM + LSTM para detección de régimen
# y predicción de dirección/magnitud de precio
# ============================================================

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import yfinance as yf
from datetime import datetime, timedelta
from tqdm import tqdm

# HMM
from hmmlearn.hmm import GaussianHMM

# LSTM 
import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

# Métricas
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import accuracy_score

# Excel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ====================
# Activar modo interactivo
# ====================
plt.ion()

# ============================================================
# PARÁMETROS CONFIGURABLES
# ============================================================
TICKER          = input("Ingresa el ticker a analizar (ej. AAPL, TSLA, AMZN): ").strip().upper()
START_DATE      = "2010-01-01"
END_DATE        = "2026-06-01"
N_PRED_DIAS     = 5       # días a predecir hacia adelante
VENTANA_LSTM    = 60      # días de historia que ve el LSTM
MAX_REGIMENES   = 4       # BIC probará de 2 hasta este número
EPOCHS          = 200
BATCH_SIZE      = 32
RF_ANUAL        = 0.04    # tasa libre de riesgo anual
UMBRAL_COMPRA   = 0.60    # confianza mínima para señal de compra
UMBRAL_VENTA    = 0.60    # confianza mínima para señal de venta

# Ruta de salida — misma carpeta que el script
RUTA_SALIDA = os.path.dirname(os.path.abspath(__file__))

# Cripto: opera 7 días, mercado tradicional: días hábiles
CRYPTO_SUFFIXES = ("-USD", "-BTC", "-ETH", "-USDT")
ES_CRIPTO = any(TICKER.upper().endswith(s) for s in CRYPTO_SUFFIXES)

print(f"\n{'='*55}")
print(f"  HMM Híbrido — {TICKER}")
print(f"  Período: {START_DATE} → {END_DATE}")
print(f"  Tipo: {'Cripto 24/7' if ES_CRIPTO else 'Mercado tradicional'}")
print(f"{'='*55}\n")


# ============================================================
# DESCARGA DE DATOS
# ============================================================
print("📥 Descargando datos históricos...")

data = yf.download(TICKER, start=START_DATE, end=END_DATE,
                   auto_adjust=False, progress=False)
if data.empty:
    raise ValueError(f"No se encontraron datos para {TICKER}.")

# VIX como modificador global
data_vix = yf.download("^VIX", start=START_DATE, end=END_DATE,
                        auto_adjust=False, progress=False)

# Alinear VIX con los datos del ticker
vix_series = data_vix['Close'].reindex(data.index).ffill().bfill()
vix_actual  = float(vix_series.iloc[-1])

print(f"✅ Datos descargados: {len(data)} días")
print(f"📡 VIX actual: {vix_actual:.2f}\n")

# ============================================================
# CONSTRUCCIÓN DE FEATURES
# ============================================================
print("⚙️  Calculando features técnicas...")

# Extraer series base como arrays 1D limpios
precios  = pd.Series(np.asarray(data['Close'].values,  dtype=float).flatten(),
                     index=data.index)
volumenes = pd.Series(np.asarray(data['Volume'].values, dtype=float).flatten(),
                      index=data.index)
vix_s    = pd.Series(np.asarray(vix_series.values,     dtype=float).flatten(),
                     index=data.index)

# --- Retornos ---
ret_1d  = precios.pct_change()                          # retorno diario
ret_5d  = precios.pct_change(5)                         # retorno acumulado 5 días
ret_20d = precios.pct_change(20)                        # retorno acumulado 20 días

# --- Volatilidad móvil ---
vol_20d = ret_1d.rolling(20).std()                      # volatilidad 20 días

# --- RSI 14 ---
def calcular_rsi(serie, periodo=14):
    delta  = serie.diff()
    ganancia = delta.clip(lower=0).rolling(periodo).mean()
    perdida  = (-delta.clip(upper=0)).rolling(periodo).mean()
    rs  = ganancia / perdida.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return rsi / 100.0                                  # normalizado 0-1

rsi = calcular_rsi(precios)

# --- MACD normalizado ---
ema_12   = precios.ewm(span=12, adjust=False).mean()
ema_26   = precios.ewm(span=26, adjust=False).mean()
macd_raw = ema_12 - ema_26
# Normalizar por precio para que sea comparable entre tickers
macd_norm = macd_raw / precios

# --- Volumen ratio ---
vol_ratio = volumenes / volumenes.rolling(20).mean()

# --- VIX normalizado ---
vix_norm = vix_s / 100.0

# --- Consolidar DataFrame de features ---
df_feat = pd.DataFrame({
    'precio'    : precios,
    'ret_1d'    : ret_1d,
    'ret_5d'    : ret_5d,
    'ret_20d'   : ret_20d,
    'vol_20d'   : vol_20d,
    'rsi'       : rsi,
    'macd_norm' : macd_norm,
    'vol_ratio' : vol_ratio,
    'vix_norm'  : vix_norm,
}, index=data.index)

# Eliminar filas con NaN (primeros ~26 días por MACD/vol)
df_feat.dropna(inplace=True)

# Columnas que entran al HMM y al LSTM (sin precio raw)
FEATURE_COLS = ['ret_1d', 'ret_5d', 'ret_20d',
                'vol_20d', 'rsi', 'macd_norm',
                'vol_ratio', 'vix_norm']

# ============================================================
# SPLIT TEMPORAL BASE
# ============================================================

split_hmm = int(len(df_feat) * 0.80)

df_train_hmm = df_feat.iloc[:split_hmm].copy()
df_val_hmm   = df_feat.iloc[split_hmm:].copy()


print(f"✅ Features calculadas: {len(df_feat)} filas × {len(FEATURE_COLS)} features")
print(f"   Período efectivo: {df_feat.index[0].date()} → {df_feat.index[-1].date()}\n")

# ============================================================
# HMM — SELECCIÓN AUTOMÁTICA DE REGÍMENES POR BIC
# ============================================================
print("🔍 Entrenando HMM y seleccionando número óptimo de regímenes...")

# Escalar features antes del HMM
scaler_hmm = StandardScaler()

X_hmm_train = scaler_hmm.fit_transform(
    df_train_hmm[FEATURE_COLS].values
)

X_hmm_val = scaler_hmm.transform(
    df_val_hmm[FEATURE_COLS].values
)

X_hmm = np.vstack([
    X_hmm_train,
    X_hmm_val
])
# --- Búsqueda por BIC ---
# BIC = -2 * log-likelihood + k * log(n)
# k = número de parámetros del modelo
# Menor BIC = mejor balance ajuste/complejidad

resultados_bic = []

for n in tqdm(range(2, MAX_REGIMENES + 1),
              desc="Entrenando HMM",
              unit="régimen"):
    try:
        modelo = GaussianHMM(
            n_components=n,
            covariance_type="full", #cambiado de full  # cada régimen tiene su propia matriz de covarianza
            n_iter=200,
            random_state=42,
            tol=1e-3
        )
        
        modelo.fit(X_hmm_train)

        log_likelihood = modelo.score(X_hmm_train)
        
        # Número de parámetros:
        # - matriz de transición: n*(n-1)
        # - medias: n * n_features
        # - covarianzas full: n * n_features * (n_features + 1) / 2

        n_feat = X_hmm.shape[1]
        k = (n - 1) + (n * (n - 1)) + (n * n_feat) + (n * n_feat * (n_feat + 1) // 2)
        # (n-1)              -> distribución inicial (startprob_)
        # n*(n-1)            -> matriz de transición
        # n*n_feat           -> medias
        # n*n_feat*(n_feat+1)/2 -> covarianzas full

        bic = -2 * log_likelihood + k * np.log(len(X_hmm))

        resultados_bic.append((n, bic, modelo, log_likelihood))
        print(f"   Regímenes: {n}  |  BIC: {bic:,.1f}  |  LogL: {log_likelihood:,.1f}")

    except Exception as e:
        print(f"   ⚠️ n={n} falló: {e}")
        continue

if not resultados_bic:
    raise RuntimeError("El HMM no pudo ajustarse con ningún número de regímenes.")

# Seleccionar el modelo con menor BIC
resultados_bic.sort(key=lambda x: x[1])
N_OPTIMO, bic_optimo, modelo_hmm, _ = resultados_bic[0]

print(f"\n✅ Número óptimo de regímenes: {N_OPTIMO}  (BIC: {bic_optimo:,.1f})\n")

# --- Decodificar secuencia de regímenes ---
estados_raw = modelo_hmm.predict(X_hmm)              # régimen por día (0, 1, 2...)
probs_estado = modelo_hmm.predict_proba(X_hmm)       # probabilidades por día

# --- Caracterizar cada régimen automáticamente ---
# Para nombrar cada régimen analizamos sus estadísticas promedio
df_feat['regimen_raw'] = estados_raw

estadisticas_regimen = {}
for r in range(N_OPTIMO):
    mask = df_feat['regimen_raw'] == r
    stats = {
        'ret_medio'  : df_feat.loc[mask, 'ret_1d'].mean(),
        'vol_media'  : df_feat.loc[mask, 'vol_20d'].mean(),
        'rsi_medio'  : df_feat.loc[mask, 'rsi'].mean(),
        'vix_medio'  : df_feat.loc[mask, 'vix_norm'].mean() * 100,
        'n_dias'     : mask.sum()
    }
    estadisticas_regimen[r] = stats

# --- Asignar nombre descriptivo a cada régimen ---
def asignar_nombre_regimen(stats, todos_stats):
    """
    Asigna nombre según retorno medio y volatilidad relativa
    comparando contra la media global de todos los regímenes.
    """
    vol_global  = np.mean([s['vol_media']  for s in todos_stats.values()])
    ret_global  = np.mean([s['ret_medio']  for s in todos_stats.values()])
    vix_global  = np.mean([s['vix_medio']  for s in todos_stats.values()])

    es_alta_vol = stats['vol_media']  > vol_global * 1.3
    es_baja_vol = stats['vol_media']  < vol_global * 0.8
    es_alcista  = stats['ret_medio']  > ret_global + 0.0002
    es_bajista  = stats['ret_medio']  < ret_global - 0.0002
    es_alto_vix = stats['vix_medio']  > vix_global * 1.2

    if es_alta_vol and es_bajista and es_alto_vix:
        return "🔴 Pánico Bajista"
    elif es_alta_vol and es_bajista:
        return "🟠 Corrección"
    elif es_alta_vol and es_alcista:
        return "🟡 Recuperación Volátil"
    elif es_baja_vol and es_alcista:
        return "🟢 Calma Alcista"
    elif es_baja_vol and es_bajista:
        return "🔵 Deterioro Silencioso"
    else:
        return "⚪ Lateralización"

nombres_regimen = {}
for r, stats in estadisticas_regimen.items():
    nombre = asignar_nombre_regimen(stats, estadisticas_regimen)
    # Si dos regímenes quedan con el mismo nombre, agregar sufijo
    conteo = sum(1 for n in nombres_regimen.values() if n == nombre)
    if conteo > 0:
        nombre = f"{nombre} ({conteo + 1})"
    nombres_regimen[r] = nombre

# Mapear nombres al DataFrame
df_feat['regimen_nombre'] = df_feat['regimen_raw'].map(nombres_regimen)

# --- Imprimir resumen de regímenes ---
print("📊 Resumen de regímenes detectados:")
print(f"{'─'*58}")
for r in range(N_OPTIMO):
    s = estadisticas_regimen[r]
    print(f"  [{r}] {nombres_regimen[r]:<28} "
          f"días: {s['n_dias']:>4}  "
          f"ret: {s['ret_medio']:+.4f}  "
          f"vol: {s['vol_media']:.4f}  "
          f"VIX: {s['vix_medio']:.1f}")
print(f"{'─'*58}")

# Régimen actual (último día)
regimen_actual_idx  = int(estados_raw[-1])
regimen_actual_nom  = nombres_regimen[regimen_actual_idx]
probs_actual        = probs_estado[-1]

print(f"\n📍 Régimen actual: {regimen_actual_nom}")
print(f"   Probabilidades: " +
      " | ".join([f"{nombres_regimen[i].split()[1][:6]}: {probs_actual[i]:.1%}"
                  for i in range(N_OPTIMO)]))

# --- Matriz de transición ---
trans_matrix = modelo_hmm.transmat_
print(f"\n🔄 Probabilidad de permanecer en régimen actual: "
      f"{trans_matrix[regimen_actual_idx, regimen_actual_idx]:.1%}\n")

# ============================================================
# PREPARACIÓN DE SECUENCIAS PARA EL LSTM
# ============================================================
print("🧠 Preparando secuencias para el LSTM...")

# Scaler independiente para el LSTM (no reusar el del HMM)
scaler_lstm = StandardScaler()
X_scaled    = scaler_lstm.fit_transform(df_feat[FEATURE_COLS].values)

# One-hot encoding del régimen HMM → columnas adicionales al input
regimen_onehot = np.zeros((len(estados_raw), N_OPTIMO), dtype=np.float32)
for i, r in enumerate(estados_raw):
    regimen_onehot[i, r] = 1.0

# Input completo = features escaladas + régimen one-hot
X_completo = np.concatenate([X_scaled, regimen_onehot], axis=1).astype(np.float32)

# --- Etiquetas ---

precios_arr = df_feat['precio'].values

etiquetas_dir = np.full(len(precios_arr), -1, dtype=np.int64)  # -1 = no calculado (relleno)

for i in range(len(precios_arr) - N_PRED_DIAS):
    cambio = (precios_arr[i + N_PRED_DIAS] - precios_arr[i]) / precios_arr[i]
    etiquetas_dir[i] = 1 if cambio > 0 else 0   # 1=COMPRA, 0=VENTA

# Magnitud: retorno % real en N_PRED_DIAS (target de regresión)
etiquetas_mag = np.zeros(len(precios_arr), dtype=np.float32)
for i in range(len(precios_arr) - N_PRED_DIAS):
    etiquetas_mag[i] = abs((precios_arr[i + N_PRED_DIAS] - precios_arr[i]) / precios_arr[i])
# --- Construir ventanas deslizantes ---
# Cada muestra = VENTANA_LSTM días de historia → predice día VENTANA_LSTM+N_PRED_DIAS
X_seq, y_dir, y_mag = [], [], []

for i in range(VENTANA_LSTM, len(X_completo) - N_PRED_DIAS):
    X_seq.append(X_completo[i - VENTANA_LSTM : i])   # ventana de 30 días
    y_dir.append(etiquetas_dir[i])
    y_mag.append(etiquetas_mag[i])

X_seq  = np.array(X_seq,  dtype=np.float32)
y_dir  = np.array(y_dir,  dtype=np.int64)
y_mag  = np.array(y_mag,  dtype=np.float32)

# Verificación: ninguna etiqueta de relleno debe haber entrado al dataset
assert (y_dir != -1).all(), "BUG: hay etiquetas de relleno (-1) dentro de las secuencias de entrenamiento"

print(f"   Secuencias generadas : {len(X_seq)}")
print(f"   Shape input LSTM     : {X_seq.shape}  "
      f"[muestras × ventana × features+régimen]")
print(f"Distribución clases  : "
      f"VENTA={np.sum(y_dir==0)}  "
      f"COMPRA={np.sum(y_dir==1)}\n")

# --- Split 80/20 cronológico (sin shuffle — serie de tiempo) ---
split     = int(len(X_seq) * 0.80)
X_train   = torch.tensor(X_seq[:split])
X_val     = torch.tensor(X_seq[split:])
y_dir_tr  = torch.tensor(y_dir[:split])
y_dir_val = torch.tensor(y_dir[split:])
y_mag_tr  = torch.tensor(y_mag[:split]).unsqueeze(1)
y_mag_val = torch.tensor(y_mag[split:]).unsqueeze(1)

# DataLoaders
train_ds = TensorDataset(X_train, y_dir_tr, y_mag_tr)
val_ds   = TensorDataset(X_val,   y_dir_val, y_mag_val)
train_dl = DataLoader(train_ds, batch_size=BATCH_SIZE, shuffle=True)
val_dl   = DataLoader(val_ds,   batch_size=BATCH_SIZE, shuffle=False)

# ============================================================
# ARQUITECTURA LSTM HÍBRIDO
# ============================================================
class LSTMHibrido(nn.Module):
    def __init__(self, n_features, n_regimenes, hidden=128, hidden2=64):
        super().__init__()
        input_size = n_features + n_regimenes

        self.lstm1   = nn.LSTM(input_size, hidden, batch_first=True,
                                num_layers=2, dropout=0.3)
        self.bnorm1  = nn.BatchNorm1d(hidden)
        self.drop1   = nn.Dropout(0.3)
        self.dense1  = nn.Linear(hidden, hidden2)
        self.bnorm2  = nn.BatchNorm1d(hidden2)
        self.relu    = nn.ReLU()
        self.drop2   = nn.Dropout(0.2)

        # Cabeza clasificación — dirección
        self.cabeza_dir = nn.Sequential(
            nn.Linear(hidden2, 16),
            nn.ReLU(),
            nn.Linear(16, 2)
        )

        # Cabeza regresión — magnitud
        self.cabeza_mag = nn.Sequential(
            nn.Linear(hidden2, 16),
            nn.ReLU(),
            nn.Linear(16, 1)
        )

    def forward(self, x):
        out, _ = self.lstm1(x)
        out     = out[:, -1, :]           # último paso temporal
        out     = self.bnorm1(out)
        out     = self.drop1(out)
        out     = self.relu(self.dense1(out))
        out     = self.bnorm2(out)
        out     = self.drop2(out)
        dir_out = self.cabeza_dir(out)
        mag_out = self.cabeza_mag(out)
        return dir_out, mag_out
# ============================================================
# ENTRENAMIENTO CON EARLY STOPPING
# ============================================================
print("🏋️  Entrenando LSTM híbrido...")

N_FEATURES_LSTM = len(FEATURE_COLS)
modelo_lstm     = LSTMHibrido(N_FEATURES_LSTM, N_OPTIMO)

# Pesos de clase para compensar desbalance (NEUTRO suele dominar)
conteos   = np.bincount(y_dir[:split])
pesos_cls = torch.tensor(1.0 / (conteos + 1e-6), dtype=torch.float32)
pesos_cls = pesos_cls / pesos_cls.sum() * 2   # normalizar a 2 clases

criterio_dir = nn.CrossEntropyLoss(weight=pesos_cls)
criterio_mag = nn.MSELoss()
optimizador  = torch.optim.Adam(modelo_lstm.parameters(), lr=1e-4)
scheduler    = torch.optim.lr_scheduler.ReduceLROnPlateau(
                    optimizador, patience=5, factor=0.5)

# Early stopping
PACIENCIA      = 25
mejor_val_loss = float('inf')
epocas_sin_mej = 0
mejor_estado   = None

historial = {'train_loss': [], 'val_loss': [], 'val_acc': []}

barra_epocas = tqdm(range(EPOCHS), desc="Entrenando LSTM", unit="época")
for epoca in barra_epocas:
    # --- Entrenamiento ---
    modelo_lstm.train()
    train_loss = 0.0
    for xb, yb_dir, yb_mag in train_dl:
        optimizador.zero_grad()
        pred_dir, pred_mag = modelo_lstm(xb)
        loss = criterio_dir(pred_dir, yb_dir) + 0.3 * criterio_mag(pred_mag, yb_mag)
        loss.backward()
        nn.utils.clip_grad_norm_(modelo_lstm.parameters(), 1.0)  # evitar exploding gradients
        optimizador.step()
        train_loss += loss.item()
    train_loss /= len(train_dl)

    # --- Validación ---
    modelo_lstm.eval()
    val_loss = 0.0
    preds_dir_all, true_dir_all = [], []

    with torch.no_grad():
        for xb, yb_dir, yb_mag in val_dl:
            pred_dir, pred_mag = modelo_lstm(xb)
            loss = criterio_dir(pred_dir, yb_dir) + 0.3 * criterio_mag(pred_mag, yb_mag)
            val_loss += loss.item()
            preds_dir_all.extend(pred_dir.argmax(dim=1).numpy())
            true_dir_all.extend(yb_dir.numpy())
    val_loss /= len(val_dl)

    acc = accuracy_score(true_dir_all, preds_dir_all)
    scheduler.step(val_loss)

    historial['train_loss'].append(train_loss)
    historial['val_loss'].append(val_loss)
    historial['val_acc'].append(acc)

    # Imprimir cada 10 épocas
    barra_epocas.set_postfix({
    'train': f"{train_loss:.4f}",
    'val':   f"{val_loss:.4f}",
    'acc':   f"{acc:.2%}"
})

    # Early stopping
    if val_loss < mejor_val_loss - 1e-4:
        mejor_val_loss = val_loss
        epocas_sin_mej = 0
        mejor_estado   = {k: v.clone() for k, v in modelo_lstm.state_dict().items()}
    else:
        epocas_sin_mej += 1
        if epocas_sin_mej >= PACIENCIA:
            print(f"\n   ⏹️  Early stopping en época {epoca+1} "
                  f"(sin mejora por {PACIENCIA} épocas)")
            break

# Restaurar mejor modelo
if mejor_estado:
    modelo_lstm.load_state_dict(mejor_estado)

print(f"\n✅ Entrenamiento completado")
print(f"   Mejor val loss : {mejor_val_loss:.4f}")
print(f"   Val accuracy   : {max(historial['val_acc']):.2%}\n")

# ============================================================
# PREDICCIÓN CON EL MODELO ENTRENADO
# ============================================================
print("🔮 Generando predicciones...")

modelo_lstm.eval()

# --- Predicción sobre validación (out-of-sample) ---
preds_dir_val, preds_mag_val, probs_dir_val = [], [], []

with torch.no_grad():
    for xb, _, _ in val_dl:
        pred_dir, pred_mag = modelo_lstm(xb)
        probs   = torch.softmax(pred_dir, dim=1)
        preds_dir_val.extend(probs.argmax(dim=1).numpy())
        probs_dir_val.extend(probs.numpy())
        preds_mag_val.extend(pred_mag.squeeze().numpy())

preds_dir_val = np.array(preds_dir_val)
preds_mag_val = np.array(preds_mag_val)
probs_dir_val = np.array(probs_dir_val)

acc_val = accuracy_score(y_dir[split:len(y_dir)], preds_dir_val)

# --- Predicción actual (última ventana disponible) ---
ultima_ventana = torch.tensor(
    X_completo[-VENTANA_LSTM:].reshape(1, VENTANA_LSTM, -1)
)

with torch.no_grad():
    pred_dir_hoy, pred_mag_hoy = modelo_lstm(ultima_ventana)
    probs_hoy   = torch.softmax(pred_dir_hoy, dim=1).numpy()[0]
    mag_hoy     = float(pred_mag_hoy.squeeze().numpy())

dir_hoy         = int(np.argmax(probs_hoy))
confianza_hoy   = float(probs_hoy[dir_hoy])

ETIQUETAS_DIR   = {0: "VENTA", 1: "COMPRA"}
dir_hoy_txt     = ETIQUETAS_DIR[dir_hoy]

print(f"   Predicción dirección : {dir_hoy_txt}  "
      f"(confianza: {confianza_hoy:.1%})")
print(f"   Magnitud esperada    : {mag_hoy:+.2%} en {N_PRED_DIAS} días")
print(f"   Accuracy validación  : {acc_val:.2%}\n")

# Proyección precio 5 días
precio_actual = float(df_feat['precio'].iloc[-1])
mag_diaria    = mag_hoy / N_PRED_DIAS
signo         = 1 if dir_hoy == 1 else -1

proyeccion_dias = []
print(f"  📅 Proyección {N_PRED_DIAS} días hábiles:")
fecha_iter = datetime.now()
for d in range(1, N_PRED_DIAS + 1):
    fecha_iter += timedelta(days=1)
    while fecha_iter.weekday() >= 5:
        fecha_iter += timedelta(days=1)
    precio_est = precio_actual * (1 + signo * mag_diaria * d)
    proyeccion_dias.append((fecha_iter.strftime('%Y-%m-%d'), precio_est,
                             signo * mag_diaria * d))
    print(f"     {fecha_iter.strftime('%Y-%m-%d')}  "
          f"${precio_est:.2f}  ({signo * mag_diaria * d:+.2%})")

# ============================================================
# SEÑAL FINAL CONSOLIDADA
# ============================================================
def generar_senal_final(dir_txt, confianza, mag, regimen_nom,
                        trans_prob_actual, vix_actual,
                        umbral_compra, umbral_venta):
    """
    Combina régimen HMM + predicción LSTM + VIX
    para generar una señal final con nivel de confianza ajustado.
    """
    # --- Factor VIX ---
    if vix_actual >= 35:
        factor_vix   = 0.60    # pánico extremo — penalizar compras fuertemente
        contexto_vix = "⚠️  VIX EXTREMO"
    elif vix_actual >= 25:
        factor_vix   = 0.80    # miedo moderado
        contexto_vix = "🟡 VIX elevado"
    else:
        factor_vix   = 1.00    # mercado tranquilo
        contexto_vix = "🟢 VIX normal"

    # --- Factor régimen ---
    # Regímenes bajistas penalizan compras, alcistas las favorecen
    nom_lower = regimen_nom.lower()
    if "pánico" in nom_lower or "crisis" in nom_lower:
        factor_reg = 0.65
    elif "corrección" in nom_lower or "deterioro" in nom_lower:
        factor_reg = 0.80
    elif "lateralización" in nom_lower:
        factor_reg = 0.90
    elif "recuperación" in nom_lower:
        factor_reg = 0.95
    else:                      # calma alcista u otros alcistas
        factor_reg = 1.05

    # Confianza ajustada por VIX y régimen
    # Solo penaliza señales de COMPRA — las de VENTA no se penalizan
    if dir_txt == "COMPRA":
        confianza_ajustada = confianza * factor_vix * factor_reg
    elif dir_txt == "VENTA":
        # En pánico las ventas se refuerzan ligeramente
        confianza_ajustada = confianza * (2.0 - factor_reg)
        confianza_ajustada = min(confianza_ajustada, 0.99)
    else:
        confianza_ajustada = confianza

    # Estabilidad del régimen actual
    estabilidad = f"{trans_prob_actual:.1%} prob. de permanecer"

    # --- Señal final ---
    if dir_txt == "COMPRA" and confianza_ajustada >= umbral_compra:
        if confianza_ajustada >= 0.80:
            senal     = "★ FUERTE COMPRA"
            icono     = "🔥"
            nivel     = "ALTO"
        else:
            senal     = "✅ COMPRA MODERADA"
            icono     = "✅"
            nivel     = "MEDIO"

    elif dir_txt == "VENTA" and confianza_ajustada >= umbral_venta:
        if confianza_ajustada >= 0.80:
            senal     = "★ FUERTE VENTA"
            icono     = "💀"
            nivel     = "ALTO"
        else:
            senal     = "🔴 VENTA MODERADA"
            icono     = "🔴"
            nivel     = "MEDIO"

    elif dir_txt == "NEUTRO":
        senal         = "⚪ NEUTRO"
        icono         = "⚪"
        nivel         = "BAJO"

    else:
        # Dirección de compra/venta pero confianza insuficiente tras ajuste
        senal         = "🟡 SEÑAL DÉBIL"
        icono         = "🟡"
        nivel         = "BAJO"

    return {
        'senal'              : senal,
        'icono'              : icono,
        'nivel'              : nivel,
        'dir_raw'            : dir_txt,
        'confianza_raw'      : confianza,
        'confianza_ajustada' : confianza_ajustada,
        'mag_esperada'       : mag,
        'factor_vix'         : factor_vix,
        'factor_reg'         : factor_reg,
        'contexto_vix'       : contexto_vix,
        'estabilidad_reg'    : estabilidad,
    }

resultado = generar_senal_final(
    dir_hoy_txt,
    confianza_hoy,
    mag_hoy,
    regimen_actual_nom,
    trans_matrix[regimen_actual_idx, regimen_actual_idx],
    vix_actual,
    UMBRAL_COMPRA,
    UMBRAL_VENTA
)

# ============================================================
# RESUMEN EN CONSOLA
# ============================================================
SEP = "=" * 58

print(SEP)
print(f"  {resultado['icono']}  SEÑAL FINAL HMM HÍBRIDO — {TICKER}")
print(SEP)
print(f"  Señal            : {resultado['senal']}")
print(f"  Nivel confianza  : {resultado['nivel']}")
print(f"  Dirección LSTM   : {resultado['dir_raw']}  "
      f"(conf. raw: {resultado['confianza_raw']:.1%})")
print(f"  Conf. ajustada   : {resultado['confianza_ajustada']:.1%}  "
      f"(tras VIX + régimen)")
print(f"  Magnitud esperada: {resultado['mag_esperada']:+.2%} "
      f"en {N_PRED_DIAS} días")
print(f"  Régimen actual   : {regimen_actual_nom}")
print(f"  Estabilidad reg. : {resultado['estabilidad_reg']}")
print(f"  Contexto VIX     : {resultado['contexto_vix']}  ({vix_actual:.1f})")
print(f"  Factor VIX       : {resultado['factor_vix']:.2f}  |  "
      f"Factor régimen: {resultado['factor_reg']:.2f}")
print(f"  Accuracy val.    : {acc_val:.2%}  "
      f"(sobre {len(preds_dir_val)} muestras out-of-sample)")
print(SEP)

# Probabilidades por dirección hoy
for i, lbl in enumerate(["VENTA", "COMPRA"]):
    barra = "█" * int(probs_hoy[i] * 30)
    print(f"    {lbl:<8}: {probs_hoy[i]:.1%}  {barra}")

# Probabilidades de régimen hoy
print(f"\n  Probabilidades de régimen hoy:")
for i in range(N_OPTIMO):
    barra = "█" * int(probs_actual[i] * 30)
    print(f"    {nombres_regimen[i]:<30}: {probs_actual[i]:.1%}  {barra}")

print(f"\n{SEP}\n")

# ============================================================
# GRÁFICAS
# ============================================================
print("📊 Generando gráficas...")

# Paleta de colores por régimen — consistente en todas las gráficas
PALETA = [
    '#2ecc71', '#3498db', '#f39c12',
    '#e74c3c', '#9b59b6', '#1abc9c'
]
colores_regimen = {r: PALETA[r % len(PALETA)] for r in range(N_OPTIMO)}

# Fechas del período efectivo (post-dropna)
fechas = df_feat.index
n_compras = np.sum(preds_dir_val == 1)
n_ventas  = np.sum(preds_dir_val == 0)
print(f"   Señales validación: COMPRA={n_compras} ({n_compras/len(preds_dir_val):.1%})  "
      f"VENTA={n_ventas} ({n_ventas/len(preds_dir_val):.1%})")
fig, axes = plt.subplots(2, 2, figsize=(18, 12))
fig.suptitle(f"HMM Híbrido — {TICKER}   |   "
             f"Régimen actual: {regimen_actual_nom}   |   "
             f"Señal: {resultado['senal']}",
             fontsize=13, fontweight='bold', y=0.98)

# ── Gráfica 1: Precio con regímenes coloreados de fondo ──────
ax1 = axes[0, 0]

# Sombrear fondo según régimen
# Suavizar régimen con ventana móvil para reducir ruido visual
regimen_suav = df_feat['regimen_raw'].rolling(15, min_periods=1).median().astype(int)
regimen_arr  = regimen_suav.values
en_regimen   = None
inicio_bloque = None

for i in range(len(fechas)):
    reg_i = int(regimen_arr[i])
    if reg_i != en_regimen:
        if en_regimen is not None:
            ax1.axvspan(inicio_bloque, fechas[i],
                        color=colores_regimen[en_regimen], alpha=0.25)
        en_regimen    = reg_i
        inicio_bloque = fechas[i]
# Cerrar último bloque
if en_regimen is not None:
    ax1.axvspan(inicio_bloque, fechas[-1],
                color=colores_regimen[en_regimen], alpha=0.25)

ax1.plot(fechas, df_feat['precio'].values,
         color='white' if plt.rcParams['axes.facecolor'] == '#1a1a2e'
         else '#2c3e50',
         linewidth=1.2, label='Precio')

# Señales LSTM sobre validación (últimos split días)
fechas_val = fechas[VENTANA_LSTM + split : VENTANA_LSTM + split + len(preds_dir_val)]
precios_val = df_feat['precio'].values[VENTANA_LSTM + split :
                                        VENTANA_LSTM + split + len(preds_dir_val)]

for i, (f, p, d) in enumerate(zip(fechas_val, precios_val, preds_dir_val)):
    if d == 1:   # COMPRA
        ax1.scatter(f, p, marker='^', color='#2ecc71', s=40, zorder=5, alpha=0.7)
    elif d == 0: # VENTA
        ax1.scatter(f, p, marker='v', color='#e74c3c', s=40, zorder=5, alpha=0.7)

# Leyenda regímenes
parches = [mpatches.Patch(color=colores_regimen[r], alpha=0.5,
                           label=nombres_regimen[r])
           for r in range(N_OPTIMO)]
parches += [mpatches.Patch(color='#2ecc71', label='▲ Señal COMPRA (val.)'),
            mpatches.Patch(color='#e74c3c', label='▼ Señal VENTA (val.)'),
            mpatches.Patch(color='#f39c12', label='◆ Proyección 5d')]
ax1.legend(handles=parches, fontsize=7, loc='upper left')
ax1.set_title('Precio + Regímenes HMM + Señales LSTM', fontsize=10)
ax1.set_ylabel('Precio')
ax1.grid(True, alpha=0.3)
ax1.tick_params(axis='x', rotation=30)

# Proyección futura en gráfica 1
fechas_proy   = [pd.Timestamp(p[0]) for p in proyeccion_dias]
precios_proy  = [p[1] for p in proyeccion_dias]
ax1.plot(fechas_proy, precios_proy, 'o--', color='#f39c12',
         linewidth=1.5, markersize=5, label='Proyección 5d', zorder=6)
for fp, pp, delta in zip(fechas_proy, precios_proy, proyeccion_dias):
    ax1.annotate(f"${pp:.1f}", (fp, pp),
                 textcoords="offset points", xytext=(0, 8),
                 ha='center', fontsize=7, color='#f39c12')

# ── Gráfica 2: Probabilidades de régimen apiladas ────────────
ax2 = axes[0, 1]

# Suavizar probabilidades con media móvil de 10 días para legibilidad
prob_suav = pd.DataFrame(probs_estado,
                          index=fechas,
                          columns=[nombres_regimen[r] for r in range(N_OPTIMO)])
prob_suav = prob_suav.rolling(60, min_periods=1).mean()

ax2.stackplot(fechas,
              [prob_suav.iloc[:, r].values for r in range(N_OPTIMO)],
              labels=[nombres_regimen[r] for r in range(N_OPTIMO)],
              colors=[PALETA[r % len(PALETA)] for r in range(N_OPTIMO)],
              alpha=0.75)
ax2.set_title('Probabilidades de Régimen (suavizadas 10d)', fontsize=10)
ax2.set_ylabel('Probabilidad')
ax2.set_ylim(0, 1)
ax2.legend(fontsize=7, loc='upper left')
ax2.grid(True, alpha=0.3)
ax2.tick_params(axis='x', rotation=30)

# ── Gráfica 3: Magnitud predicha vs real (out-of-sample) ─────
ax3 = axes[1, 0]

mag_real_val = etiquetas_mag[split : split + len(preds_mag_val)]
fechas_mag   = fechas[VENTANA_LSTM + split :
                       VENTANA_LSTM + split + len(preds_mag_val)]

ax3.plot(fechas_mag, mag_real_val * 100,
         label='Magnitud real (%)', color='#3498db', linewidth=1.0, alpha=0.8)
ax3.plot(fechas_mag, preds_mag_val * 100,
         label='Magnitud predicha (%)', color='#e67e22',
         linewidth=1.0, linestyle='--', alpha=0.8)
ax3.axhline(0, color='gray', linewidth=0.8, linestyle=':')
ax3.fill_between(fechas_mag,
                  mag_real_val * 100,
                  preds_mag_val * 100,
                  alpha=0.15, color='#e74c3c', label='Error')

# Métricas de regresión
mae_mag  = np.mean(np.abs(mag_real_val - preds_mag_val)) * 100
rmse_mag = np.sqrt(np.mean((mag_real_val - preds_mag_val) ** 2)) * 100
ax3.set_title(f'Magnitud Predicha vs Real (val.)  '
              f'MAE={mae_mag:.3f}%  RMSE={rmse_mag:.3f}%', fontsize=10)
ax3.set_ylabel('Retorno esperado (%)')
ax3.legend(fontsize=8)
ax3.grid(True, alpha=0.3)
ax3.tick_params(axis='x', rotation=30)

# ── Gráfica 4: Matriz de transición entre regímenes ──────────
ax4 = axes[1, 1]

im = ax4.imshow(trans_matrix, cmap='YlOrRd', vmin=0, vmax=1)
plt.colorbar(im, ax=ax4, shrink=0.8)

etiquetas_cortas = [nombres_regimen[r].split()[1][:8]
                    if len(nombres_regimen[r].split()) > 1
                    else nombres_regimen[r][:8]
                    for r in range(N_OPTIMO)]

ax4.set_xticks(range(N_OPTIMO))
ax4.set_yticks(range(N_OPTIMO))
ax4.set_xticklabels(etiquetas_cortas, rotation=30, ha='right', fontsize=8)
ax4.set_yticklabels(etiquetas_cortas, fontsize=8)
ax4.set_title('Matriz de Transición entre Regímenes', fontsize=10)
ax4.set_xlabel('Régimen siguiente')
ax4.set_ylabel('Régimen actual')

# Anotar probabilidades en cada celda
for i in range(N_OPTIMO):
    for j in range(N_OPTIMO):
        val   = trans_matrix[i, j]
        color = 'white' if val > 0.6 else 'black'
        ax4.text(j, i, f'{val:.2f}',
                 ha='center', va='center',
                 fontsize=9, color=color, fontweight='bold')

# Resaltar diagonal (probabilidad de permanecer)
for i in range(N_OPTIMO):
    ax4.add_patch(plt.Rectangle((i - 0.5, i - 0.5), 1, 1,
                                  fill=False, edgecolor='#2c3e50',
                                  linewidth=2.5))

plt.tight_layout()
plt.savefig(os.path.join(RUTA_SALIDA,
            f"HMM_Hibrido_{TICKER}_{datetime.now().strftime('%Y%m%d_%H%M')}.png"),
            dpi=150, bbox_inches='tight')
plt.show()
print("✅ Gráficas generadas\n")

# ============================================================
# EXPORTACIÓN A EXCEL
# ============================================================
print("💾 Exportando resultados a Excel...")

nombre_excel = f"HMM_Hibrido_{TICKER}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
ruta_excel   = os.path.join(RUTA_SALIDA, nombre_excel)

wb = Workbook()

# --- Estilos reutilizables ---
header_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
header_font  = Font(color="FFFFFF", bold=True)
borde        = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)
centro       = Alignment(horizontal='center')

fill_compra       = PatternFill(start_color="00C851", end_color="00C851", fill_type="solid")
fill_venta        = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
fill_neutro       = PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid")
fill_senal_fuerte = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")

def escribir_encabezado(ws, headers, ancho=18):
    """Escribe fila de encabezado con formato estándar."""
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = centro
        cell.border    = borde
        ws.column_dimensions[get_column_letter(col)].width = ancho

def aplicar_formato_fila(ws, fila, valores, fill=None):
    """Escribe una fila con borde, centrado y fill opcional."""
    for col, val in enumerate(valores, 1):
        cell           = ws.cell(row=fila, column=col, value=val)
        cell.border    = borde
        cell.alignment = centro
        if fill:
            cell.fill  = fill

# ── Hoja 1: Resumen ejecutivo ─────────────────────────────────
ws_res = wb.active
ws_res.title = "Resumen"

escribir_encabezado(ws_res, ['Campo', 'Valor'], ancho=35)

resumen_filas = [
    ("Ticker",                  TICKER),
    ("Fecha análisis",          datetime.now().strftime('%Y-%m-%d %H:%M')),
    ("Período datos",           f"{START_DATE} → {END_DATE}"),
    ("─── SEÑAL FINAL ───",     ""),
    ("Señal",                   resultado['senal']),
    ("Nivel confianza",         resultado['nivel']),
    ("Dirección LSTM",          resultado['dir_raw']),
    ("Confianza raw",           f"{resultado['confianza_raw']:.2%}"),
    ("Confianza ajustada",      f"{resultado['confianza_ajustada']:.2%}"),
    ("Magnitud esperada",       f"{resultado['mag_esperada']:+.2%} en {N_PRED_DIAS} días"),
    ("─── PROYECCIÓN 5 DÍAS ───", ""),
] + [
    (f"Día {i+1} — {proyeccion_dias[i][0]}",
     f"${proyeccion_dias[i][1]:.2f}  ({proyeccion_dias[i][2]:+.2%})")
    for i in range(len(proyeccion_dias))
] + [
    ("─── RÉGIMEN HMM ───",     ""),
    ("Regímenes óptimos (BIC)", N_OPTIMO),
    ("Régimen actual",          regimen_actual_nom),
    ("Prob. permanecer",        f"{trans_matrix[regimen_actual_idx, regimen_actual_idx]:.1%}"),
    ("─── FACTORES ───",        ""),
    ("VIX actual",              f"{vix_actual:.2f}"),
    ("Contexto VIX",            resultado['contexto_vix']),
    ("Factor VIX",              f"{resultado['factor_vix']:.2f}"),
    ("Factor régimen",          f"{resultado['factor_reg']:.2f}"),
    ("─── MODELO ───",          ""),
    ("Accuracy validación",     f"{acc_val:.2%}"),
    ("MAE magnitud (val.)",     f"{mae_mag:.4f}%"),
    ("RMSE magnitud (val.)",    f"{rmse_mag:.4f}%"),
    ("Ventana LSTM",            f"{VENTANA_LSTM} días"),
    ("Épocas entrenadas",       len(historial['train_loss'])),
    ("Mejor val loss",          f"{mejor_val_loss:.6f}"),
]

for fila_idx, (campo, valor) in enumerate(resumen_filas, start=2):
    es_separador = str(campo).startswith("───")
    fill_fila    = PatternFill(start_color="2D4A7A", end_color="2D4A7A",
                                fill_type="solid") if es_separador else None
    font_fila    = Font(color="FFFFFF", bold=True) if es_separador else None

    for col in [1, 2]:
        val  = campo if col == 1 else (valor if not es_separador else "")
        cell = ws_res.cell(row=fila_idx, column=col, value=val)
        cell.border    = borde
        cell.alignment = centro
        if fill_fila:
            cell.fill  = fill_fila
        if font_fila:
            cell.font  = font_fila

    # Colorear la fila de señal final
    if campo == "Señal":
        color_senal = ("00C851" if "COMPRA" in str(valor)
                        else "D32F2F" if "VENTA" in str(valor)
                        else "9E9E9E")
        for col in [1, 2]:
            ws_res.cell(row=fila_idx, column=col).fill = PatternFill(
                start_color=color_senal, end_color=color_senal, fill_type="solid")
            ws_res.cell(row=fila_idx, column=col).font = Font(
                color="FFFFFF", bold=True)

# ── Hoja 2: Regímenes por día ─────────────────────────────────
ws_reg = wb.create_sheet("Regímenes")

headers_reg = ['Fecha', 'Régimen', 'Nombre Régimen',
               'Ret 1d (%)', 'Vol 20d', 'RSI', 'VIX'] + \
              [f'Prob {nombres_regimen[r].split()[1][:8]}'
               for r in range(N_OPTIMO)]

escribir_encabezado(ws_reg, headers_reg, ancho=16)

# Colores para cada régimen en la hoja
fill_regimenes = {r: PatternFill(
    start_color=PALETA[r % len(PALETA)].replace('#', ''),
    end_color=PALETA[r % len(PALETA)].replace('#', ''),
    fill_type="solid") for r in range(N_OPTIMO)}

for fila_idx, (fecha, row) in enumerate(df_feat.iterrows(), start=2):
    reg_idx = int(row['regimen_raw'])
    valores = [
        fecha.strftime('%Y-%m-%d'),
        reg_idx,
        nombres_regimen[reg_idx],
        f"{row['ret_1d'] * 100:+.3f}",
        f"{row['vol_20d']:.5f}",
        f"{row['rsi']:.3f}",
        f"{row['vix_norm'] * 100:.1f}",
    ] + [f"{probs_estado[fila_idx - 2, r]:.3f}" for r in range(N_OPTIMO)]

    aplicar_formato_fila(ws_reg, fila_idx, valores,
                          fill=fill_regimenes[reg_idx])

# ── Hoja 3: Predicciones out-of-sample ───────────────────────
ws_pred = wb.create_sheet("Predicciones")

headers_pred = ['Fecha', 'Precio Real', 'Dirección Pred.',
                'Dirección Real', 'Correcta', 'Prob Venta',
                'Prob Compra',
                'Magnitud Pred. (%)', 'Magnitud Real (%)']
escribir_encabezado(ws_pred, headers_pred, ancho=18)

dir_real_val   = y_dir[split : split + len(preds_dir_val)]
fechas_pred    = fechas[VENTANA_LSTM + split :
                         VENTANA_LSTM + split + len(preds_dir_val)]
precios_pred   = df_feat['precio'].values[VENTANA_LSTM + split :
                                           VENTANA_LSTM + split + len(preds_dir_val)]

for fila_idx, i in enumerate(range(len(preds_dir_val)), start=2):
    dir_pred_txt = ETIQUETAS_DIR[int(preds_dir_val[i])]
    dir_real_txt = ETIQUETAS_DIR[int(dir_real_val[i])]
    correcta     = "✅" if preds_dir_val[i] == dir_real_val[i] else "❌"

    valores = [
        fechas_pred[i].strftime('%Y-%m-%d'),
        f"{precios_pred[i]:.2f}",
        dir_pred_txt,
        dir_real_txt,
        correcta,
        f"{probs_dir_val[i, 0]:.3f}",
        f"{probs_dir_val[i, 1]:.3f}",
        
        f"{preds_mag_val[i] * 100:+.3f}",
        f"{mag_real_val[i] * 100:+.3f}",
    ]

    # Color por dirección predicha
    if dir_pred_txt == "COMPRA":
        fill_fila = fill_compra
    elif dir_pred_txt == "VENTA":
        fill_fila = fill_venta
    else:
        fill_fila = fill_neutro

    aplicar_formato_fila(ws_pred, fila_idx, valores, fill=fill_fila)

# ── Hoja 4: Matriz de transición ─────────────────────────────
ws_trans = wb.create_sheet("Transición")

# Encabezado con nombres de régimen
headers_trans = ['Régimen origen \\ destino'] + \
                [nombres_regimen[r] for r in range(N_OPTIMO)]
escribir_encabezado(ws_trans, headers_trans, ancho=22)

for i in range(N_OPTIMO):
    fila_vals = [nombres_regimen[i]] + \
                [f"{trans_matrix[i, j]:.4f}" for j in range(N_OPTIMO)]
    for col, val in enumerate(fila_vals, 1):
        cell           = ws_trans.cell(row=i + 2, column=col, value=val)
        cell.border    = borde
        cell.alignment = centro
        # Colorear diagonal (permanecer en régimen)
        if col == i + 2:
            intensidad = int(trans_matrix[i, i] * 200) + 55
            hex_color  = f"{intensidad:02X}FF{intensidad:02X}"
            cell.fill  = PatternFill(start_color=hex_color,
                                      end_color=hex_color,
                                      fill_type="solid")
        # Colorear primera columna como encabezado lateral
        if col == 1:
            cell.fill  = fill_regimenes[i]
            cell.font  = Font(color="FFFFFF", bold=True)

wb.save(ruta_excel)
print(f"✅ Excel guardado: {nombre_excel}")
print(f"   📄 Hoja 1 — Resumen ejecutivo")
print(f"   📄 Hoja 2 — Regímenes por día ({len(df_feat)} filas)")
print(f"   📄 Hoja 3 — Predicciones out-of-sample ({len(preds_dir_val)} filas)")
print(f"   📄 Hoja 4 — Matriz de transición")
# ============================================================
# BLOQUE 8 — VALIDACIÓN EXTENDIDA
# ============================================================
# UBICACIÓN: pegar justo después de wb.save(ruta_excel) y antes del mensaje de cierre final

print("\n📋 Validación extendida...")

from sklearn.metrics import confusion_matrix, classification_report

# ── Validación 1: Confusion matrix global ────────────────────
print("\n1️⃣  Confusion Matrix global (validación):")
dir_real_v8 = y_dir[split : split + len(preds_dir_val)]
cm = confusion_matrix(dir_real_v8, preds_dir_val)
print(f"   {'':12} {'Pred VENTA':>12} {'Pred COMPRA':>12}")
for i, lbl in enumerate(["Real VENTA ", "Real COMPRA"]):
    print(f"   {lbl:12} {cm[i,0]:>12} {cm[i,1]:>12}")

report = classification_report(dir_real_v8, preds_dir_val,
                                target_names=["VENTA", "COMPRA"],
                                output_dict=True)
print(f"\n   Precisión VENTA  : {report['VENTA']['precision']:.2%}")
print(f"   Recall    VENTA  : {report['VENTA']['recall']:.2%}")
print(f"   Precisión COMPRA : {report['COMPRA']['precision']:.2%}")
print(f"   Recall    COMPRA : {report['COMPRA']['recall']:.2%}")

# ── Validación 2: Accuracy por régimen ───────────────────────
print("\n2️⃣  Accuracy por régimen (validación):")
print(f"   {'Régimen':<30} {'Días':>6} {'Aciertos':>9} {'Accuracy':>10}")
print(f"   {'─'*58}")

idx_v8_inicio = VENTANA_LSTM + split
idx_v8_fin    = VENTANA_LSTM + split + len(preds_dir_val)
regimen_v8    = df_feat['regimen_raw'].values[idx_v8_inicio : idx_v8_fin]
fechas_v8     = fechas[idx_v8_inicio : idx_v8_fin]
precios_v8    = df_feat['precio'].values[idx_v8_inicio : idx_v8_fin]

acc_por_regimen = {}
for r in range(N_OPTIMO):
    mask_r = regimen_v8 == r
    if mask_r.sum() < 10:
        continue
    acc_r    = accuracy_score(dir_real_v8[mask_r], preds_dir_val[mask_r])
    acc_por_regimen[r] = acc_r
    aciertos = int(acc_r * mask_r.sum())
    print(f"   {nombres_regimen[r]:<30} {mask_r.sum():>6} {aciertos:>9} {acc_r:>9.2%}")

# ── Validación 3: Accuracy por año ───────────────────────────
print("\n3️⃣  Accuracy por año (validación):")
print(f"   {'Año':>6} {'Días':>6} {'Aciertos':>9} {'Accuracy':>10}")
print(f"   {'─'*35}")

anios = sorted(set(f.year for f in fechas_v8))
acc_por_anio = {}
for anio in anios:
    mask_a = np.array([f.year == anio for f in fechas_v8])
    if mask_a.sum() < 10:
        continue
    acc_a    = accuracy_score(dir_real_v8[mask_a], preds_dir_val[mask_a])
    acc_por_anio[anio] = acc_a
    aciertos = int(acc_a * mask_a.sum())
    print(f"   {anio:>6} {mask_a.sum():>6} {aciertos:>9} {acc_a:>9.2%}")

# ── Validación 4: Sharpe ratio simulado ──────────────────────
print("\n4️⃣  Sharpe ratio simulado (validación):")

retornos_diarios    = np.diff(precios_v8) / precios_v8[:-1]
posicion            = np.where(preds_dir_val[:-1] == 1, 1, 0).astype(float)
retornos_estrategia = posicion * retornos_diarios
retornos_buyhold    = retornos_diarios
rf_diario           = RF_ANUAL / 252

sharpe_estrategia = ((retornos_estrategia.mean() - rf_diario) /
                      (retornos_estrategia.std() + 1e-8) * np.sqrt(252))
sharpe_buyhold    = ((retornos_buyhold.mean() - rf_diario) /
                      (retornos_buyhold.std() + 1e-8) * np.sqrt(252))

retorno_acum_est  = (1 + retornos_estrategia).cumprod()[-1] - 1
retorno_acum_bh   = (1 + retornos_buyhold).cumprod()[-1] - 1

equity_curve = (1 + retornos_estrategia).cumprod()
rolling_max  = np.maximum.accumulate(equity_curve)
drawdowns    = (equity_curve - rolling_max) / rolling_max
max_dd       = drawdowns.min()

print(f"   {'Métrica':<28} {'Estrategia':>12} {'Buy & Hold':>12}")
print(f"   {'─'*54}")
print(f"   {'Sharpe Ratio':<28} {sharpe_estrategia:>12.3f} {sharpe_buyhold:>12.3f}")
print(f"   {'Retorno acumulado':<28} {retorno_acum_est:>11.2%} {retorno_acum_bh:>11.2%}")
print(f"   {'Máx. Drawdown':<28} {max_dd:>11.2%} {'—':>12}")
print(f"   {'Días en posición larga':<28} {int(np.sum(posicion==1)):>12} {'—':>12}")
print(f"   {'Días en cash (fuera)':<28} {int(np.sum(posicion==0)):>12} {'—':>12}")
# ── Gráfica de validación extendida ──────────────────────────
fig_val, axes_val = plt.subplots(1, 3, figsize=(18, 5))
fig_val.suptitle(f"Validación Extendida — {TICKER}", fontsize=13, fontweight='bold')

# Subplot 1: Equity curve
ax_eq  = axes_val[0]
equity_bh = (1 + retornos_buyhold).cumprod()
ax_eq.plot(fechas_v8[1:], equity_curve,
           label=f'Estrategia HMM ({retorno_acum_est:+.1%})',
           color='#2ecc71', linewidth=1.5)
ax_eq.plot(fechas_v8[1:], equity_bh,
           label=f'Buy & Hold ({retorno_acum_bh:+.1%})',
           color='#3498db', linewidth=1.5, linestyle='--')
ax_eq.axhline(1, color='gray', linewidth=0.8, linestyle=':')
ax_eq.set_title('Equity Curve (validación)', fontsize=10)
ax_eq.set_ylabel('Valor relativo')
ax_eq.legend(fontsize=8)
ax_eq.grid(True, alpha=0.3)
ax_eq.tick_params(axis='x', rotation=30)

# Subplot 2: Accuracy por régimen
ax_reg = axes_val[1]
if acc_por_regimen:
    nombres_r = [nombres_regimen[r].split()[1][:10] for r in acc_por_regimen]
    valores_r = list(acc_por_regimen.values())
    colores_r = ['#2ecc71' if v >= 0.55 else '#e74c3c' if v < 0.50
                  else '#f39c12' for v in valores_r]
    bars = ax_reg.bar(nombres_r, valores_r, color=colores_r, alpha=0.8)
    ax_reg.axhline(0.5,     color='gray',    linewidth=1, linestyle='--', label='Azar (50%)')
    ax_reg.axhline(acc_val, color='#3498db', linewidth=1, linestyle='-.', 
                   label=f'Acc. global ({acc_val:.1%})')
    for bar, val in zip(bars, valores_r):
        ax_reg.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.01,
                    f'{val:.1%}', ha='center', fontsize=8)
    ax_reg.set_title('Accuracy por Régimen', fontsize=10)
    ax_reg.set_ylabel('Accuracy')
    ax_reg.set_ylim(0, 1)
    ax_reg.legend(fontsize=8)
    ax_reg.grid(True, alpha=0.3, axis='y')
    ax_reg.tick_params(axis='x', rotation=30)

# Subplot 3: Accuracy por año
ax_anio = axes_val[2]
if acc_por_anio:
    anios_str = [str(a) for a in acc_por_anio]
    valores_a = list(acc_por_anio.values())
    colores_a = ['#2ecc71' if v >= 0.55 else '#e74c3c' if v < 0.50
                  else '#f39c12' for v in valores_a]
    bars2 = ax_anio.bar(anios_str, valores_a, color=colores_a, alpha=0.8)
    ax_anio.axhline(0.5,     color='gray',    linewidth=1, linestyle='--', label='Azar (50%)')
    ax_anio.axhline(acc_val, color='#3498db', linewidth=1, linestyle='-.',
                    label=f'Acc. global ({acc_val:.1%})')
    for bar, val in zip(bars2, valores_a):
        ax_anio.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.01,
                     f'{val:.1%}', ha='center', fontsize=8)
    ax_anio.set_title('Accuracy por Año', fontsize=10)
    ax_anio.set_ylabel('Accuracy')
    ax_anio.set_ylim(0, 1)
    ax_anio.legend(fontsize=8)
    ax_anio.grid(True, alpha=0.3, axis='y')

plt.tight_layout()
plt.savefig(os.path.join(RUTA_SALIDA,
            f"HMM_Validacion_{TICKER}_{datetime.now().strftime('%Y%m%d_%H%M')}.png"),
            dpi=150, bbox_inches='tight')
plt.show()
print("\n✅ Validación extendida completada\n")

print(f"\n{'='*55}")
print(f"  ✅ Script completado — {TICKER}")
print(f"  📁 Archivos guardados en: {RUTA_SALIDA}")
print(f"{'='*55}\n")

# Fin del script
plt.ioff()
plt.show()
