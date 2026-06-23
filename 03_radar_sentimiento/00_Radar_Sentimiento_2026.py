import tkinter as tk
import customtkinter as ctk
import threading
import requests
import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from textblob import TextBlob
from pytrends.request import TrendReq
from datetime import datetime, timedelta
import pytz
import json
import time
import os
import warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
warnings.filterwarnings('ignore')

# =============================
# CONFIGURACIÓN
# =============================
FINNHUB_API_KEY = "d8iubupr01qmeauk8hsgd8iubupr01qmeauk8ht0"

TICKERS_BASE = [
    'AAPL', 'AMZN', 'GOOG', 'NVDA', 'TSLA', 'META', 'MSFT',
    'AVGO', 'CAT', 'LLY', 'JPM', 'WMT', 'XOM', 'MA', 'UNH',
    'PLTR', 'MARA', 'GME', 'SOFI', 'UBER', 'AMD', 'INTC',
    'NFLX', 'ORCL', 'MRK', 'WFC', 'FCX', 'FSLR', 'CLF', 'GE'
]

INTERVALO_ABIERTO  = 300 #tiempo de refresco cuando el mercado está abierto o en pre/post-market
INTERVALO_CERRADO  = 1800 #tiempo de inactividad cuando el mercado está cerrado
TOP_N_VIRAL        = 30  #máximo de tickers virales a analizar por ciclo
SCORE_MINIMO       = 0.01 #score mínimo absoluto para mostrar en resultados
ET_ZONE            = pytz.timezone('America/New_York') #zona horaria para cálculos de mercado

# Mapa de exchanges conocidos
EXCHANGE_MAP = {
    'NYSE': '🇺🇸 NYSE', 'NMS': '🇺🇸 NASDAQ', 'NGM': '🇺🇸 NASDAQ',
    'PCX': '🇺🇸 NYSE', 'ASE': '🇺🇸 AMEX', 'MEX': '🇲🇽 BMV',
    'LSE': '🇬🇧 LSE', 'TSE': '🇯🇵 Tokyo', 'HKG': '🇭🇰 HK',
    'SHH': '🇨🇳 Shanghai', 'FRA': '🇩🇪 Frankfurt', 'EPA': '🇫🇷 Paris',
}

# =============================
# ESTADO DEL MERCADO
# =============================
def get_market_status():
    now_et = datetime.now(ET_ZONE)
    weekday = now_et.weekday()
    hora = now_et.hour + now_et.minute / 60
    if weekday >= 5:
        return "CERRADO", "🔴", INTERVALO_CERRADO
    if 4.0 <= hora < 9.5:
        return "PRE-MARKET", "🟡", INTERVALO_ABIERTO
    elif 9.5 <= hora < 16.0:
        return "ABIERTO", "🟢", INTERVALO_ABIERTO
    elif 16.0 <= hora < 20.0:
        return "AFTER-HOURS", "🔵", INTERVALO_ABIERTO
    else:
        return "CERRADO", "🔴", INTERVALO_CERRADO

def get_mercados_internacionales():
    """Verifica si mercados asiáticos y europeos están abiertos y su tendencia."""
    now_utc = datetime.utcnow()
    hora_utc = now_utc.hour + now_utc.minute / 60
    resultados = {}

    # Tokyo: UTC+9, abre 9am-3pm JST = 0:00-6:00 UTC
    tokyo_abierto = 0.0 <= hora_utc < 6.0 or hora_utc >= 23.0
    resultados['tokyo'] = ('🟢 Abierto' if tokyo_abierto else '🔴 Cerrado', tokyo_abierto)

    # Shanghai: UTC+8, abre 9:30am-3pm CST = 1:30-7:00 UTC
    shanghai_abierto = 1.5 <= hora_utc < 7.0
    resultados['shanghai'] = ('🟢 Abierto' if shanghai_abierto else '🔴 Cerrado', shanghai_abierto)

    # Londres: UTC+1 (BST), abre 8am-4:30pm = 7:00-15:30 UTC
    london_abierto = 7.0 <= hora_utc < 15.5
    resultados['london'] = ('🟢 Abierto' if london_abierto else '🔴 Cerrado', london_abierto)

    # Frankfurt: UTC+2 (CEST), abre 9am-5:30pm = 7:00-15:30 UTC
    frankfurt_abierto = 7.0 <= hora_utc < 15.5
    resultados['frankfurt'] = ('🟢 Abierto' if frankfurt_abierto else '🔴 Cerrado', frankfurt_abierto)

    return resultados

def get_spillover_asiatico():
    """
    Detecta el efecto spillover asiático.
    Si Asia cerró en rojo, ajusta el score negativamente.
    """
    try:
        # Índices asiáticos como proxy
        indices = {
            'Nikkei': '^N225',
            'Hang Seng': '^HSI',
            'Shanghai': '000001.SS'
        }
        spillover = 0.0
        count = 0
        for nombre, sym in indices.items():
            try:
                t = yf.Ticker(sym)
                hist = t.history(period='2d')
                if len(hist) >= 2:
                    ret = (hist['Close'].iloc[-1] - hist['Close'].iloc[-2]) / hist['Close'].iloc[-2]
                    spillover += ret
                    count += 1
            except:
                pass
        if count > 0:
            avg = spillover / count
            return float(np.clip(avg * 5, -0.3, 0.3))
        return 0.0
    except:
        return 0.0

# =============================
# INFO DE EMPRESA
# =============================
_cache_empresas = {}

def get_company_info(ticker):
    """Obtiene nombre de empresa y exchange via yfinance con caché."""
    if ticker in _cache_empresas:
        return _cache_empresas[ticker]
    try:
        t = yf.Ticker(ticker)
        info = t.info
        nombre = info.get('shortName') or info.get('longName') or ticker
        exchange_raw = info.get('exchange', '')
        exchange = EXCHANGE_MAP.get(exchange_raw, f'🌐 {exchange_raw}' if exchange_raw else '🌐 --')
        sector = info.get('sector', '--')
        resultado = {'nombre': nombre, 'exchange': exchange, 'sector': sector}
        _cache_empresas[ticker] = resultado
        return resultado
    except:
        resultado = {'nombre': ticker, 'exchange': '🌐 --', 'sector': '--'}
        _cache_empresas[ticker] = resultado
        return resultado

# =============================
# FUENTES DE DATOS
# =============================
def get_fear_greed():
    try:
        url = "https://production.dataviz.cnn.io/index/fearandgreed/graphdata"
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=10)
        data = r.json()
        score = data['fear_and_greed']['score']
        rating = data['fear_and_greed']['rating']
        return float(score), rating
    except:
        return 50.0, "Neutral"

def get_finnhub_sentiment(ticker):
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        url = f"https://finnhub.io/api/v1/company-news?symbol={ticker}&from={week_ago}&to={today}&token={FINNHUB_API_KEY}"
        r = requests.get(url, timeout=10)
        noticias = r.json()
        if not noticias or not isinstance(noticias, list):
            return 0.0, 0
        scores = []
        for n in noticias[:10]:
            texto = f"{n.get('headline', '')} {n.get('summary', '')}"
            if texto.strip():
                blob = TextBlob(texto)
                scores.append(blob.sentiment.polarity)
        return np.mean(scores) if scores else 0.0, len(noticias)
    except:
        return 0.0, 0

def get_stocktwits_sentiment(ticker):
    try:
        url = f"https://api.stocktwits.com/api/2/streams/symbol/{ticker}.json"
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=10)
        data = r.json()
        messages = data.get('messages', [])
        if not messages:
            return 0.0, 0
        bullish = sum(1 for m in messages if m.get('entities', {}).get('sentiment', {}) and
                     m['entities']['sentiment'].get('basic') == 'Bullish')
        bearish = sum(1 for m in messages if m.get('entities', {}).get('sentiment', {}) and
                     m['entities']['sentiment'].get('basic') == 'Bearish')
        total = bullish + bearish
        if total == 0:
            scores = []
            for m in messages[:20]:
                texto = m.get('body', '')
                if texto:
                    blob = TextBlob(texto)
                    scores.append(blob.sentiment.polarity)
            return np.mean(scores) if scores else 0.0, len(messages)
        return (bullish - bearish) / total, len(messages)
    except:
        return 0.0, 0

def get_yahoo_news_sentiment(ticker):
    try:
        t = yf.Ticker(ticker)
        noticias = t.news
        if not noticias:
            return 0.0, 0
        scores = []
        for n in noticias[:10]:
            titulo = n.get('title', '')
            if titulo:
                blob = TextBlob(titulo)
                scores.append(blob.sentiment.polarity)
        return np.mean(scores) if scores else 0.0, len(noticias)
    except:
        return 0.0, 0

def get_insider_signal(ticker):
    try:
        url = f"https://efts.sec.gov/LATEST/search-index?q=%22{ticker}%22&dateRange=custom&startdt={(datetime.now()-timedelta(days=30)).strftime('%Y-%m-%d')}&enddt={datetime.now().strftime('%Y-%m-%d')}&forms=4"
        headers = {"User-Agent": "radar@actinver.com"}
        r = requests.get(url, headers=headers, timeout=10)
        data = r.json()
        hits = data.get('hits', {}).get('hits', [])
        compras = sum(1 for h in hits[:20] if h.get('_source', {}).get('form_type') == '4')
        if compras > 3:
            return 0.5, compras
        elif compras > 0:
            return 0.2, compras
        return 0.0, 0
    except:
        return 0.0, 0

def get_price_momentum(ticker):
    try:
        t = yf.Ticker(ticker)
        hist = t.history(period="1mo")
        if hist.empty or len(hist) < 5:
            return 0.0, 0.0
        ret_semana = (hist['Close'].iloc[-1] - hist['Close'].iloc[-5]) / hist['Close'].iloc[-5]
        ret_mes    = (hist['Close'].iloc[-1] - hist['Close'].iloc[0])  / hist['Close'].iloc[0]
        vol_reciente = hist['Volume'].iloc[-5:].mean()
        vol_historico = hist['Volume'].mean()
        vol_ratio = vol_reciente / vol_historico if vol_historico > 0 else 1.0
        mom_score = np.clip(ret_semana * 3 + ret_mes, -1, 1)
        return float(mom_score), float(vol_ratio)
    except:
        return 0.0, 1.0

def get_google_trends(tickers_batch):
    try:
        pytrends = TrendReq(hl='en-US', tz=360)
        batch = tickers_batch[:5]
        pytrends.build_payload(batch, cat=0, timeframe='now 7-d', geo='US')
        df = pytrends.interest_over_time()
        if df.empty:
            return {t: 0.0 for t in batch}
        scores = {}
        for t in batch:
            if t in df.columns:
                serie = df[t].values
                if len(serie) > 1:
                    trend = (serie[-1] - serie[0]) / (serie[0] + 1)
                    scores[t] = float(np.clip(trend, -1, 1))
                else:
                    scores[t] = 0.0
            else:
                scores[t] = 0.0
        return scores
    except:
        return {t: 0.0 for t in tickers_batch[:5]}

# =============================
# SCORE COMPUESTO
# =============================
def calcular_score(ticker, fear_greed_score, google_trends_scores, spillover=0.0):
    resultados = {'ticker': ticker}

    finn_score, finn_count = get_finnhub_sentiment(ticker)
    st_score, st_count     = get_stocktwits_sentiment(ticker)
    yh_score, yh_count     = get_yahoo_news_sentiment(ticker)
    ins_score, ins_count   = get_insider_signal(ticker)
    mom_score, vol_ratio   = get_price_momentum(ticker)
    gt_score               = google_trends_scores.get(ticker, 0.0)

    # Limpiar todos los valores
    finn_score = 0.0  if (finn_score  is None or np.isnan(float(finn_score  if finn_score  is not None else 0))) else float(finn_score)
    st_score   = 0.0  if (st_score    is None or np.isnan(float(st_score    if st_score    is not None else 0))) else float(st_score)
    yh_score   = 0.0  if (yh_score    is None or np.isnan(float(yh_score    if yh_score    is not None else 0))) else float(yh_score)
    ins_score  = 0.0  if (ins_score   is None or np.isnan(float(ins_score   if ins_score   is not None else 0))) else float(ins_score)
    mom_score  = 0.0  if (mom_score   is None or np.isnan(float(mom_score   if mom_score   is not None else 0))) else float(mom_score)
    vol_ratio  = 1.0  if (vol_ratio   is None or np.isnan(float(vol_ratio   if vol_ratio   is not None else 1))) else float(vol_ratio)
    gt_score   = 0.0  if (gt_score    is None or np.isnan(float(gt_score    if gt_score    is not None else 0))) else float(gt_score)
    spillover  = 0.0  if (spillover   is None or np.isnan(float(spillover   if spillover   is not None else 0))) else float(spillover)

    resultados['finnhub_score']        = finn_score
    resultados['finnhub_noticias']     = finn_count
    resultados['stocktwits_score']     = st_score
    resultados['stocktwits_mensajes']  = st_count
    resultados['yahoo_score']          = yh_score
    resultados['yahoo_noticias']       = yh_count
    resultados['insider_score']        = ins_score
    resultados['insider_transacciones']= ins_count
    resultados['momentum_score']       = mom_score
    resultados['volumen_ratio']        = vol_ratio
    resultados['google_score']         = gt_score
    resultados['spillover']            = spillover

    score = (
        finn_score * 0.25 +
        st_score   * 0.20 +
        yh_score   * 0.15 +
        ins_score  * 0.15 +
        mom_score  * 0.15 +
        gt_score   * 0.10
    )
    score += spillover * 0.1

    if fear_greed_score < 25:
        score *= 0.7
    elif fear_greed_score > 75:
        score *= 0.85

    if vol_ratio > 1.5:
        score *= 1.15

    resultados['score_final'] = float(np.clip(score, -1, 1))

    info = get_company_info(ticker)
    resultados['nombre']   = info['nombre']
    resultados['exchange'] = info['exchange']
    resultados['sector']   = info['sector']

    s = resultados['score_final']
    if s >= 0.5:
        resultados['señal'] = '🔥 FUERTE COMPRA'
        resultados['color'] = '#00C851'
    elif s >= 0.25:
        resultados['señal'] = '✅ COMPRA'
        resultados['color'] = '#00897B'
    elif s >= 0.05:
        resultados['señal'] = '🟡 NEUTRAL+'
        resultados['color'] = '#FFB300'
    elif s >= -0.05:
        resultados['señal'] = '⚪ NEUTRAL'
        resultados['color'] = '#9E9E9E'
    elif s >= -0.25:
        resultados['señal'] = '🟠 PRECAUCIÓN'
        resultados['color'] = '#FF6D00'
    else:
        resultados['señal'] = '🔴 EVITAR'
        resultados['color'] = '#D32F2F'

    return resultados


    # Obtener info de empresa
    info = get_company_info(ticker)
    resultados['nombre'] = info['nombre']
    resultados['exchange'] = info['exchange']
    resultados['sector'] = info['sector']

    s = resultados['score_final']
    if s >= 0.5:
        resultados['señal'] = '🔥 FUERTE COMPRA'
        resultados['color'] = '#00C851'
    elif s >= 0.25:
        resultados['señal'] = '✅ COMPRA'
        resultados['color'] = '#00897B'
    elif s >= 0.05:
        resultados['señal'] = '🟡 NEUTRAL+'
        resultados['color'] = '#FFB300'
    elif s >= -0.05:
        resultados['señal'] = '⚪ NEUTRAL'
        resultados['color'] = '#9E9E9E'
    elif s >= -0.25:
        resultados['señal'] = '🟠 PRECAUCIÓN'
        resultados['color'] = '#FF6D00'
    else:
        resultados['señal'] = '🔴 EVITAR'
        resultados['color'] = '#D32F2F'

    return resultados

def detectar_tickers_virales():
    virales = []
    try:
        url = "https://api.stocktwits.com/api/2/trending/symbols.json"
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=10)
        data = r.json()
        symbols = data.get('symbols', [])
        for s in symbols[:TOP_N_VIRAL]:
            ticker = s.get('symbol', '')
            if ticker and ticker not in TICKERS_BASE and '.' not in ticker:
                virales.append(ticker)
    except:
        pass
    return virales

# =============================
# INTERFAZ GRÁFICA
# =============================
class RadarApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Radar de Sentimiento de Mercado")
        self.geometry("1500x900")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.resultados = []
        self.escaneando = False
        self.hilo_escaneo = None
        self.fear_greed = (50.0, "Neutral")
        self.modo_escaneo = ctk.StringVar(value="Mixto")
        self._tickers_custom = list(TICKERS_BASE)

        self._build_ui()

    def _build_ui(self):
        # ── Header ──────────────────────────────────────────────
        header = ctk.CTkFrame(self, height=65, corner_radius=0)
        header.pack(fill='x', padx=0, pady=0)

        ctk.CTkLabel(header, text="📡 RADAR DE SENTIMIENTO DE MERCADO",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(side='left', padx=20, pady=10)

        self.lbl_mercado = ctk.CTkLabel(header, text="⏳ Verificando...",
                                         font=ctk.CTkFont(size=14))
        self.lbl_mercado.pack(side='left', padx=15)

        self.lbl_fg = ctk.CTkLabel(header, text="Fear & Greed: --",
                                    font=ctk.CTkFont(size=13))
        self.lbl_fg.pack(side='left', padx=15)

        # Indicadores mercados internacionales
        self.lbl_tokyo    = ctk.CTkLabel(header, text="🇯🇵 --", font=ctk.CTkFont(size=11))
        self.lbl_shanghai = ctk.CTkLabel(header, text="🇨🇳 --", font=ctk.CTkFont(size=11))
        self.lbl_london   = ctk.CTkLabel(header, text="🇬🇧 --", font=ctk.CTkFont(size=11))
        self.lbl_frankfurt= ctk.CTkLabel(header, text="🇩🇪 --", font=ctk.CTkFont(size=11))
        for lbl in [self.lbl_tokyo, self.lbl_shanghai, self.lbl_london, self.lbl_frankfurt]:
            lbl.pack(side='left', padx=8)

        self.lbl_update = ctk.CTkLabel(header, text="Última actualización: --",
                                        font=ctk.CTkFont(size=11))
        self.lbl_update.pack(side='right', padx=20)

        # ── Barra de controles ───────────────────────────────────
        ctrl = ctk.CTkFrame(self, height=55, corner_radius=0, fg_color="transparent")
        ctrl.pack(fill='x', padx=10, pady=3)

        self.btn_escanear = ctk.CTkButton(
            ctrl, text="🔍 Iniciar Escaneo", width=160,
            command=self.toggle_escaneo,
            fg_color="#1565C0", hover_color="#0D47A1"
        )
        self.btn_escanear.pack(side='left', padx=5)

        # Selector de modo
        ctk.CTkLabel(ctrl, text="Modo:", font=ctk.CTkFont(size=12)).pack(side='left', padx=(15,3))
        modo_menu = ctk.CTkOptionMenu(
            ctrl,
            values=["Favoritos", "Dinámico", "Mixto"],
            variable=self.modo_escaneo,
            width=120,
            command=self._on_modo_change
        )
        modo_menu.pack(side='left', padx=3)

        ctk.CTkButton(
            ctrl, text="✏️ Editar Lista", width=120,
            command=self._editar_lista,
            fg_color="#37474F", hover_color="#263238"
        ).pack(side='left', padx=5)

        ctk.CTkButton(
            ctrl, text="📊 Exportar Excel", width=150,
            command=self.exportar_excel,
            fg_color="#2E7D32", hover_color="#1B5E20"
        ).pack(side='left', padx=5)

        ctk.CTkButton(
            ctrl, text="📄 Exportar CSV", width=130,
            command=self.exportar_csv,
            fg_color="#6A1B9A", hover_color="#4A148C"
        ).pack(side='left', padx=5)

        # Indicador de refresco
        self.lbl_refresh = ctk.CTkLabel(ctrl, text="⬜", font=ctk.CTkFont(size=18))
        self.lbl_refresh.pack(side='left', padx=10)

        self.lbl_status = ctk.CTkLabel(ctrl, text="Listo para escanear.",
                                        font=ctk.CTkFont(size=12))
        self.lbl_status.pack(side='left', padx=5)

        # ── Panel principal ──────────────────────────────────────
        main = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        main.pack(fill='both', expand=True, padx=10, pady=5)

        # Panel izquierdo — tabla
        left = ctk.CTkFrame(main, width=560)
        left.pack(side='left', fill='both', padx=5, pady=5)
        left.pack_propagate(False)

        ctk.CTkLabel(left, text="🎯 Tickers Detectados",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=5)

        tabla_frame = ctk.CTkScrollableFrame(left)
        tabla_frame.pack(fill='both', expand=True, padx=5, pady=5)
        self.tabla_frame = tabla_frame

        # Panel derecho — gráfica
        right = ctk.CTkFrame(main)
        right.pack(side='right', fill='both', expand=True, padx=5, pady=5)

        ctk.CTkLabel(right, text="Mapa de Sentimiento",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=5)

        self.fig = Figure(figsize=(8, 6), facecolor='#1a1a2e')
        self.canvas = FigureCanvasTkAgg(self.fig, master=right)
        toolbar = NavigationToolbar2Tk(self.canvas, right)
        toolbar.update()
        toolbar.pack(side='bottom', fill='x')
        self.canvas.get_tk_widget().pack(fill='both', expand=True)

        self._init_grafica()
        self._tooltip = None
        self._annot = None
        self._scatter = None
        self._scatter_data = []
        self.canvas.mpl_connect('motion_notify_event', self._on_hover)
        self._actualizar_estado_mercado()
        self._actualizar_mercados_internacionales()

    def _on_modo_change(self, modo):
        descripciones = {
            "Favoritos": "Solo tickers de tu lista base",
            "Dinámico":  "Solo tickers virales detectados automáticamente",
            "Mixto":     "Lista base + tickers virales"
        }
        self.lbl_status.configure(text=f"Modo: {modo} — {descripciones[modo]}")

    def _editar_lista(self):
        """Ventana para editar la lista de tickers favoritos."""
        win = ctk.CTkToplevel(self)
        win.title("✏️ Editar Lista de Tickers Favoritos")
        win.geometry("400x500")
        win.grab_set()

        ctk.CTkLabel(win, text="Un ticker por línea:",
                     font=ctk.CTkFont(size=13)).pack(pady=10)

        txt = ctk.CTkTextbox(win, width=360, height=350)
        txt.pack(padx=20, pady=5)
        txt.insert("1.0", "\n".join(self._tickers_custom))

        def guardar():
            contenido = txt.get("1.0", "end").strip()
            nuevos = [t.strip().upper() for t in contenido.split('\n') if t.strip()]
            self._tickers_custom = nuevos
            self.lbl_status.configure(text=f"✅ Lista actualizada: {len(nuevos)} tickers")
            win.destroy()

        ctk.CTkButton(win, text="💾 Guardar", command=guardar,
                      fg_color="#2E7D32").pack(pady=10)

    def _init_grafica(self):
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.set_facecolor('#1a1a2e')
        ax.text(0.5, 0.5, 'Presiona "Iniciar Escaneo"\npara comenzar el análisis',
                ha='center', va='center', color='white', fontsize=14,
                transform=ax.transAxes)
        ax.set_xticks([])
        ax.set_yticks([])
        self.canvas.draw()

    def _actualizar_estado_mercado(self):
        status, icono, _ = get_market_status()
        self.lbl_mercado.configure(text=f"{icono} MERCADO: {status}")
        self.after(60000, self._actualizar_estado_mercado)

    def _actualizar_mercados_internacionales(self):
        """Actualiza indicadores de mercados internacionales."""
        mercados = get_mercados_internacionales()
        self.lbl_tokyo.configure(text=f"🇯🇵 Tokyo: {mercados['tokyo'][0]}")
        self.lbl_shanghai.configure(text=f"🇨🇳 Shanghai: {mercados['shanghai'][0]}")
        self.lbl_london.configure(text=f"🇬🇧 Londres: {mercados['london'][0]}")
        self.lbl_frankfurt.configure(text=f"🇩🇪 Frankfurt: {mercados['frankfurt'][0]}")
        self.after(300000, self._actualizar_mercados_internacionales)  # cada 5 min

    def _flash_refresh(self):
        """Parpadeo visual cuando hay actualización."""
        self.lbl_refresh.configure(text="🟢")
        self.after(500, lambda: self.lbl_refresh.configure(text="⬜"))

    def toggle_escaneo(self):
        if not self.escaneando:
            self.escaneando = True
            self.btn_escanear.configure(text="⏹ Detener Escaneo", fg_color="#C62828")
            self.hilo_escaneo = threading.Thread(target=self._ciclo_escaneo, daemon=True)
            self.hilo_escaneo.start()
        else:
            self.escaneando = False
            self.btn_escanear.configure(text="🔍 Iniciar Escaneo", fg_color="#1565C0")
            self.lbl_status.configure(text="Escaneo detenido.")

    def _ciclo_escaneo(self):
        while self.escaneando:
            self._escanear()
            _, _, intervalo = get_market_status()
            for _ in range(intervalo):
                if not self.escaneando:
                    break
                time.sleep(1)

    def _get_tickers_segun_modo(self, virales):
        modo = self.modo_escaneo.get()
        if modo == "Favoritos":
            return list(self._tickers_custom)
        elif modo == "Dinámico":
            return list(virales)
        else:  # Mixto
            return list(set(self._tickers_custom + virales))

    def _escanear(self):
        try:
            self.after(0, lambda: self.lbl_status.configure(text="🔄 Escaneando mercado..."))

            fg_score, fg_rating = get_fear_greed()
            self.fear_greed = (fg_score, fg_rating)
            self.after(0, lambda: self.lbl_fg.configure(
                text=f"Fear & Greed: {fg_score:.0f} ({fg_rating})"
            ))

            # Spillover asiático
            self.after(0, lambda: self.lbl_status.configure(text="🌏 Calculando spillover asiático..."))
            spillover = get_spillover_asiatico()

            virales = detectar_tickers_virales()
            tickers_total = self._get_tickers_segun_modo(virales)

            gt_scores = {}
            self.after(0, lambda: self.lbl_status.configure(text="📈 Consultando Google Trends..."))
            try:
                for i in range(0, min(len(tickers_total), 25), 5):
                    batch = tickers_total[i:i+5]
                    gt_batch = get_google_trends(batch)
                    gt_scores.update(gt_batch)
                    time.sleep(1)
            except:
                pass

            resultados_nuevos = []
            total = len(tickers_total)
            for idx, ticker in enumerate(tickers_total):
                if not self.escaneando:
                    break
                self.after(0, lambda t=ticker, i=idx: self.lbl_status.configure(
                    text=f" Analizando {t}... ({i+1}/{total})"
                ))
                try:
                    res = calcular_score(ticker, fg_score, gt_scores, spillover)
                    print(f"{ticker}: score={res['score_final']:.4f}")
                    if abs(res['score_final']) >= SCORE_MINIMO:
                        resultados_nuevos.append(res)
                except Exception as e:
                    import traceback
                    print(f"ERROR en {ticker}: {e}")
                    print(traceback.format_exc())
                time.sleep(0.5)

            resultados_nuevos.sort(key=lambda x: abs(x['score_final']), reverse=True)
            self.resultados = resultados_nuevos

            now = datetime.now().strftime('%H:%M:%S')
            self.after(0, lambda: self._actualizar_ui(now))

        except Exception as e:
            self.after(0, lambda: self.lbl_status.configure(text=f"⚠️ Error: {str(e)}"))

    def _actualizar_ui(self, timestamp):
        self.lbl_update.configure(text=f"Última actualización: {timestamp}")
        self.lbl_status.configure(text=f"✅ Escaneo completo — {len(self.resultados)} tickers detectados")
        self._actualizar_tabla()
        self._actualizar_grafica()
        self.after(0, self._flash_refresh)

    def _actualizar_tabla(self):
        for widget in self.tabla_frame.winfo_children():
            widget.destroy()

        headers = ['Ticker', 'Score', 'Señal', 'Momentum', 'Vol x', 'Mercado']
        anchos  = [75, 65, 135, 80, 65, 90]
        for col, (h, w) in enumerate(zip(headers, anchos)):
            ctk.CTkLabel(self.tabla_frame, text=h,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         fg_color="#1565C0", corner_radius=4,
                         width=w).grid(row=0, column=col, padx=2, pady=2, sticky='ew')

        for row_idx, res in enumerate(self.resultados[:40], start=1):
            bg = "#1e1e2e" if row_idx % 2 == 0 else "#16213e"

            ctk.CTkLabel(self.tabla_frame, text=res['ticker'],
                         font=ctk.CTkFont(size=11, weight="bold"),
                         fg_color=bg, width=75).grid(row=row_idx, column=0, padx=2, pady=1)

            score_color = "#00C851" if res['score_final'] > 0 else "#D32F2F"
            ctk.CTkLabel(self.tabla_frame, text=f"{res['score_final']:+.3f}",
                         font=ctk.CTkFont(size=11), text_color=score_color,
                         fg_color=bg, width=65).grid(row=row_idx, column=1, padx=2, pady=1)

            ctk.CTkLabel(self.tabla_frame, text=res['señal'],
                         font=ctk.CTkFont(size=10),
                         fg_color=bg, width=135).grid(row=row_idx, column=2, padx=2, pady=1)

            mom = res.get('momentum_score', 0)
            ctk.CTkLabel(self.tabla_frame, text=f"{mom:+.2f}",
                         font=ctk.CTkFont(size=11),
                         text_color="#00C851" if mom > 0 else "#D32F2F",
                         fg_color=bg, width=80).grid(row=row_idx, column=3, padx=2, pady=1)

            vol = res.get('volumen_ratio', 1.0)
            ctk.CTkLabel(self.tabla_frame, text=f"{vol:.2f}x",
                         font=ctk.CTkFont(size=11),
                         text_color="#FFB300" if vol > 1.5 else "white",
                         fg_color=bg, width=65).grid(row=row_idx, column=4, padx=2, pady=1)

            # Mercado de origen
            exchange = res.get('exchange', '🌐 --')
            ctk.CTkLabel(self.tabla_frame, text=exchange,
                         font=ctk.CTkFont(size=9),
                         fg_color=bg, width=90).grid(row=row_idx, column=5, padx=2, pady=1)

    def _actualizar_grafica(self):
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.set_facecolor('#0d1117')
        self.fig.patch.set_facecolor('#1a1a2e')

        if not self.resultados:
            self._scatter = None
            self._scatter_data = []
            ax.text(0.5, 0.5, 'Sin datos aún', ha='center', va='center',
                    color='white', fontsize=14, transform=ax.transAxes)
            self.canvas.draw()
            return

        top = self.resultados[:20]
        scores    = [r['score_final'] for r in top]
        momentos  = [r.get('momentum_score', 0) for r in top]
        volumenes = [max(r.get('volumen_ratio', 1.0) * 100, 50) for r in top]
        colores   = [r['color'] for r in top]

        scatter = ax.scatter(scores, momentos, s=volumenes,
                             c=colores, alpha=0.85, edgecolors='white',
                             linewidths=0.5, zorder=3)

        for i, r in enumerate(top):
            ax.annotate(r['ticker'], (scores[i], momentos[i]),
                        fontsize=8, color='white', ha='center', va='bottom',
                        xytext=(0, 8), textcoords='offset points')

        ax.axvline(0, color='gray', linewidth=0.8, linestyle='--', alpha=0.5)
        ax.axhline(0, color='gray', linewidth=0.8, linestyle='--', alpha=0.5)
        ax.axvline(0.25, color='#00C851', linewidth=0.5, linestyle=':', alpha=0.4)
        ax.axvline(-0.25, color='#D32F2F', linewidth=0.5, linestyle=':', alpha=0.4)
        ax.set_xlabel('Score de Sentimiento →', color='white', fontsize=10)
        ax.set_ylabel('Momentum de Precio →', color='white', fontsize=10)
        ax.set_title('Mapa de Sentimiento (tamaño = volumen inusual)',
                     color='white', fontsize=11, pad=10)
        ax.tick_params(colors='white')
        ax.spines['bottom'].set_color('gray')
        ax.spines['left'].set_color('gray')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.axvspan(0.25, 1.0, alpha=0.05, color='green')
        ax.axvspan(-1.0, -0.25, alpha=0.05, color='red')

        self._scatter = scatter
        self._scatter_data = top
        self.canvas.draw()

    def _on_hover(self, event):
        if not hasattr(self, '_scatter') or self._scatter is None or not self._scatter_data:
            return
        try:
            ax = self.fig.axes[0]
        except:
            return
        if event.inaxes is None:
            if self._annot:
                self._annot.remove()
                self._annot = None
                self.canvas.draw_idle()
            return

        cont, ind = self._scatter.contains(event)
        if cont:
            idx = ind['ind'][0]
            res = self._scatter_data[idx]
            if self._annot:
                try:
                    self._annot.remove()
                except:
                    pass

            texto = (
                f"  {res['ticker']} — {res.get('nombre', res['ticker'])}  \n"
                f"  {res.get('exchange', '🌐 --')} | {res.get('sector', '--')}  \n"
                f"  ─────────────────────────  \n"
                f"  Score:      {res['score_final']:+.3f}  \n"
                f"  Señal:      {res['señal']}  \n"
                f"  Momentum:   {res.get('momentum_score',0):+.2f}  \n"
                f"  Volumen:    {res.get('volumen_ratio',1.0):.2f}x  \n"
                f"  ─────────────────────────  \n"
                f"  Finnhub:    {res.get('finnhub_score',0):+.3f} ({res.get('finnhub_noticias',0)} noticias)  \n"
                f"  StockTwits: {res.get('stocktwits_score',0):+.3f} ({res.get('stocktwits_mensajes',0)} msgs)  \n"
                f"  Yahoo News: {res.get('yahoo_score',0):+.3f}  \n"
                f"  Insider:    {res.get('insider_score',0):+.3f}  \n"
                f"  Google:     {res.get('google_score',0):+.3f}  \n"
                f"  Spillover:  {res.get('spillover',0):+.3f}  "
            )

            self._annot = ax.annotate(
                texto,
                xy=(res['score_final'], res.get('momentum_score', 0)),
                xytext=(15, 15), textcoords='offset points',
                bbox=dict(boxstyle='round,pad=0.6', fc='#0d1117', ec='#4fc3f7', alpha=0.97, lw=1.5),
                color='white', fontsize=8.5, family='monospace',
                arrowprops=dict(arrowstyle='->', color='#4fc3f7', lw=1.2)
            )
            self.canvas.draw_idle()
        else:
            if self._annot:
                try:
                    self._annot.remove()
                except:
                    pass
                self._annot = None
                self.canvas.draw_idle()

    def exportar_excel(self):
        if not self.resultados:
            self.lbl_status.configure(text="⚠️ No hay datos para exportar.")
            return
        try:
            ruta = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            f"Radar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
            wb = Workbook()
            ws = wb.active
            ws.title = "Radar Sentimiento"

            header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

            headers = ['Ticker', 'Empresa', 'Mercado', 'Sector', 'Score Final', 'Señal',
                      'Finnhub', 'StockTwits', 'Yahoo', 'Insider', 'Momentum', 'Volumen x',
                      'Google', 'Spillover']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = 15

            fill_map = {
                '🔥 FUERTE COMPRA': PatternFill(start_color="00C851", end_color="00C851", fill_type="solid"),
                '✅ COMPRA':        PatternFill(start_color="00897B", end_color="00897B", fill_type="solid"),
                '🟡 NEUTRAL+':      PatternFill(start_color="FFB300", end_color="FFB300", fill_type="solid"),
                '⚪ NEUTRAL':       PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
                '🟠 PRECAUCIÓN':    PatternFill(start_color="FF6D00", end_color="FF6D00", fill_type="solid"),
                '🔴 EVITAR':        PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid"),
            }

            for row_idx, res in enumerate(self.resultados, start=2):
                vals = [
                    res['ticker'],
                    res.get('nombre', res['ticker']),
                    res.get('exchange', '--'),
                    res.get('sector', '--'),
                    f"{res['score_final']:+.4f}",
                    res['señal'],
                    f"{res.get('finnhub_score', 0):+.3f}",
                    f"{res.get('stocktwits_score', 0):+.3f}",
                    f"{res.get('yahoo_score', 0):+.3f}",
                    f"{res.get('insider_score', 0):+.3f}",
                    f"{res.get('momentum_score', 0):+.3f}",
                    f"{res.get('volumen_ratio', 1.0):.2f}x",
                    f"{res.get('google_score', 0):+.3f}",
                    f"{res.get('spillover', 0):+.3f}",
                ]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    if col_idx == 6:
                        cell.fill = fill_map.get(res['señal'], PatternFill())

            wb.save(ruta)
            self.lbl_status.configure(text=f"✅ Excel Generado con Exíto: {os.path.basename(ruta)}")
        except Exception as e:
            self.lbl_status.configure(text=f"⚠️ Error Excel: {str(e)}")

    def exportar_csv(self):
        if not self.resultados:
            self.lbl_status.configure(text="⚠️ No hay datos para exportar.")
            return
        try:
            ruta = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            f"Radar_Tickers_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        )
            with open(ruta, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Ticker', 'Empresa', 'Mercado', 'Score', 'Señal', 'Momentum'])
                for res in self.resultados:
                    senal_limpia = res['señal']
                    for emoji in ['🔥','✅','🟡','⚪','🟠','🔴']:
                        senal_limpia = senal_limpia.replace(emoji, '').strip()
                    writer.writerow([
                        res['ticker'],
                        res.get('nombre', res['ticker']),
                        res.get('exchange', '--'),
                        f"{res['score_final']:+.4f}",
                        senal_limpia,
                        f"{res.get('momentum_score', 0):+.3f}"
                    ])
            self.lbl_status.configure(text=f"✅ CSV Generado con Exíto: {os.path.basename(ruta)}")
        except Exception as e:
            self.lbl_status.configure(text=f"⚠️ Error CSV: {str(e)}")

# =============================
# MAIN
# =============================
if __name__ == "__main__":
    app = RadarApp()
    app.mainloop()
